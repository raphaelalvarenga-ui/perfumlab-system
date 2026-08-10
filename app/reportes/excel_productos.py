from pathlib import Path
import unicodedata

from app.controllers.productos_controller import ProductosController
from app.database.json_storage import DATABASE_PATH, cargar_tabla, es_activo
from app.models.producto import Producto
from app.validaciones import limpiar_texto, validar_ruta_exportacion


COLUMNAS_EXPORTACION = (
    ("id", "ID"),
    ("sku", "SKU"),
    ("nombre", "Nombre"),
    ("marca", "Marca"),
    ("categoria", "Categoria"),
    ("costo", "Costo"),
    ("precio", "Precio"),
    ("stock_actual", "Stock actual"),
    ("stock_minimo", "Stock minimo"),
    ("estado", "Estado"),
)

ENCABEZADOS_IMPORTACION = {
    "sku": {"sku"},
    "nombre": {"nombre", "producto"},
    "marca": {"marca"},
    "categoria": {"categoria", "categoria_id"},
    "costo": {"costo"},
    "precio": {"precio"},
    "stock_actual": {"stock_actual", "stock actual", "stock"},
    "stock_minimo": {"stock_minimo", "stock minimo"},
}

REQUERIDOS_IMPORTACION = (
    "sku",
    "nombre",
    "costo",
    "precio",
    "stock_actual",
    "stock_minimo",
)


def exportar_productos_excel(ruta_archivo, ruta_db=DATABASE_PATH):
    openpyxl = _cargar_openpyxl()
    ruta_archivo = validar_ruta_exportacion(
        ruta_archivo,
        ".xlsx",
        "el inventario Excel",
    )

    workbook = openpyxl.Workbook()
    hoja = workbook.active
    hoja.title = "Productos"
    hoja.append([encabezado for _clave, encabezado in COLUMNAS_EXPORTACION])

    for fila in obtener_filas_productos(ruta_db):
        hoja.append([fila[clave] for clave, _encabezado in COLUMNAS_EXPORTACION])

    _aplicar_formato_excel(openpyxl, hoja)
    workbook.save(ruta_archivo)
    return ruta_archivo


def importar_productos_excel(ruta_archivo, ruta_db=DATABASE_PATH):
    openpyxl = _cargar_openpyxl()
    ruta_archivo = _validar_ruta_importacion(ruta_archivo)
    workbook = openpyxl.load_workbook(ruta_archivo, read_only=True, data_only=True)
    productos = []
    errores = []

    try:
        hoja = workbook.active
        filas = hoja.iter_rows(values_only=True)
        encabezados = next(filas, None)
        if encabezados is None:
            raise ValueError("El archivo Excel esta vacio.")

        mapa_columnas = _mapear_columnas(encabezados)
        skus_archivo = {}
        skus_existentes = _obtener_skus_existentes(ruta_db)
        categorias = _obtener_categorias_por_nombre(ruta_db)

        for numero_fila, valores in enumerate(filas, start=2):
            if _fila_vacia(valores):
                continue

            datos = _extraer_datos_fila(valores, mapa_columnas)
            try:
                producto = _crear_producto_importado(datos, categorias)
                sku_normalizado = producto.sku.strip().lower()

                if sku_normalizado in skus_existentes:
                    raise ValueError(f"El SKU {producto.sku} ya esta registrado.")

                if sku_normalizado in skus_archivo:
                    raise ValueError(
                        f"El SKU {producto.sku} esta duplicado en la fila "
                        f"{skus_archivo[sku_normalizado]}."
                    )

                skus_archivo[sku_normalizado] = numero_fila
                productos.append(producto)
            except ValueError as error:
                errores.append(f"Fila {numero_fila}: {error}")
    finally:
        workbook.close()

    if errores:
        raise ValueError(_formatear_errores_importacion(errores))

    if not productos:
        raise ValueError("El archivo no contiene productos para importar.")

    ids_creados = ProductosController(ruta_db).crear_productos_lote(productos)
    return len(ids_creados)


def obtener_filas_productos(ruta_db=DATABASE_PATH):
    controlador = ProductosController(ruta_db)
    categorias = {
        int(categoria["id"]): categoria["nombre"]
        for categoria in cargar_tabla("categorias", ruta_db)
    }
    filas = []

    for producto in controlador.listar_productos(incluir_inactivos=True):
        filas.append(
            {
                "id": producto.id,
                "sku": producto.sku,
                "nombre": producto.nombre,
                "marca": producto.marca,
                "categoria": categorias.get(producto.categoria_id, ""),
                "costo": float(producto.costo),
                "precio": float(producto.precio),
                "stock_actual": int(producto.stock_actual),
                "stock_minimo": int(producto.stock_minimo),
                "estado": "Activo" if producto.activo else "Inactivo",
            }
        )

    return filas


def _crear_producto_importado(datos, categorias):
    categoria_nombre = limpiar_texto(datos.get("categoria"))
    categoria_id = None

    if categoria_nombre:
        categoria_id = categorias.get(categoria_nombre.lower())
        if categoria_id is None:
            raise ValueError(f"La categoria '{categoria_nombre}' no existe.")

    producto = Producto(
        sku=datos.get("sku"),
        nombre=datos.get("nombre"),
        categoria_id=categoria_id,
        marca=datos.get("marca"),
        costo=datos.get("costo"),
        precio=datos.get("precio"),
        stock_actual=datos.get("stock_actual"),
        stock_minimo=datos.get("stock_minimo"),
    )
    producto.validar()
    return producto


def _mapear_columnas(encabezados):
    columnas = {}

    for indice, encabezado in enumerate(encabezados):
        clave = _resolver_encabezado(encabezado)
        if clave:
            columnas[clave] = indice

    faltantes = [clave for clave in REQUERIDOS_IMPORTACION if clave not in columnas]
    if faltantes:
        nombres = ", ".join(faltantes)
        raise ValueError(f"Faltan columnas obligatorias en Excel: {nombres}.")

    return columnas


def _resolver_encabezado(encabezado):
    encabezado = _normalizar_encabezado(encabezado)

    for clave, alias in ENCABEZADOS_IMPORTACION.items():
        if encabezado in alias:
            return clave

    return None


def _normalizar_encabezado(encabezado):
    texto = limpiar_texto(encabezado).lower().replace("_", " ")
    texto = " ".join(texto.split())
    return "".join(
        caracter
        for caracter in unicodedata.normalize("NFD", texto)
        if unicodedata.category(caracter) != "Mn"
    )


def _extraer_datos_fila(valores, mapa_columnas):
    datos = {}

    for clave, indice in mapa_columnas.items():
        datos[clave] = valores[indice] if indice < len(valores) else None

    return datos


def _obtener_skus_existentes(ruta_db):
    return {
        str(producto.get("sku") or "").strip().lower()
        for producto in cargar_tabla("productos", ruta_db)
        if str(producto.get("sku") or "").strip()
    }


def _obtener_categorias_por_nombre(ruta_db):
    return {
        str(categoria.get("nombre") or "").strip().lower(): int(categoria["id"])
        for categoria in cargar_tabla("categorias", ruta_db)
        if es_activo(categoria)
    }


def _fila_vacia(valores):
    return all(limpiar_texto(valor) == "" for valor in valores)


def _validar_ruta_importacion(ruta_archivo):
    texto_ruta = limpiar_texto(ruta_archivo)

    if not texto_ruta:
        raise ValueError("Seleccione un archivo Excel para importar.")

    ruta = Path(texto_ruta)
    if not ruta.exists() or not ruta.is_file():
        raise ValueError("El archivo Excel seleccionado no existe.")

    if ruta.suffix.lower() != ".xlsx":
        raise ValueError("El archivo para importar debe tener extension .xlsx.")

    return ruta


def _aplicar_formato_excel(openpyxl, hoja):
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    relleno = PatternFill("solid", fgColor="EAF0F8")
    for celda in hoja[1]:
        celda.font = Font(bold=True)
        celda.fill = relleno

    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = hoja.dimensions

    for columna in hoja.columns:
        letra = get_column_letter(columna[0].column)
        ancho = max(len(str(celda.value or "")) for celda in columna) + 2
        hoja.column_dimensions[letra].width = min(max(ancho, 12), 28)


def _formatear_errores_importacion(errores):
    visibles = errores[:10]
    mensaje = "No se importaron productos porque hay errores:\n" + "\n".join(visibles)

    if len(errores) > len(visibles):
        mensaje += f"\n... y {len(errores) - len(visibles)} errores mas."

    return mensaje


def _cargar_openpyxl():
    try:
        import openpyxl
    except ImportError as error:
        raise RuntimeError(
            "La libreria openpyxl no esta instalada. "
            "Instalela con: python -m pip install openpyxl"
        ) from error

    return openpyxl
